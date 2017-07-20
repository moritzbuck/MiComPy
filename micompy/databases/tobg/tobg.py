from os.path import join as pjoin
import os
from tqdm import tqdm
from pandas import DataFrame, Index
import json
from micompy.common.genome import Genome
from micompy.databases.database import Database
from openpyxl import load_workbook
import shutil

class TOBG(Database) :

    def __init__(self, workbench, data_path = "/home/moritz/people/MoreData/genomes/TOBG/", clean = False):
        Database.__init__(self,workbench = workbench, data_path = data_path)

        wb = load_workbook("metadata/Table3_GenomeStats.xlsx")
        t_metadata = DataFrame([l for i,l in enumerate(wb['Sheet1'].values) if i >1], columns=[l for l in wb['Sheet1'].values][1])
        corrected = { u'\xc2Gemmatimonadetes': 'Gemmatimonadetes' ,
        'marinegroup': 'Puniceicoccaceae',
        'Urania1B19': 'Phycisphaerae',
        'Thalassopira' : 'Thalassospira',
        'SM1A02': 'Phycisphaerae',
        'SAR324cluster': 'SAR324 cluster',
        'unclassifiedAlphaproteobacteria': 'Alphaproteobacteria',
        'SAR202-2': 'SAR202 cluster',
        'SAR202-1': 'SAR202 cluster',
        'SAR116cluster' : 'SAR116 cluster',
        'OPB35soil': 'unidentified Verrucomicrobium group OPB35',
        'Pla3': 'Planctomycetes',
        'OM190': 'Planctomycetes',
        'NovelClass_B': 'Ignavibacteriae',
        'Nitropelagicus': 'Candidatus Nitrosopelagicus' ,
        'Nanoarchaoeta': 'Nanoarchaeota',
        'Methylobacterum': 'Methylobacterium',
        'JL-ENTP-F27': 'Phycisphaerae',
        'FS140-16B-02marinegroup': 'Phycisphaerae',
        'Epsilonbacteraeota': 'Bacteria',
        'DEV007': 'Verrucomicrobiales',
        'CandidatusPuniceispirillum': 'Candidatus Puniceispirillum',
        'CandidatePhylaRadiation' : 'Bacteria candidate phyla',
        'CaThioglobus': 'Candidatus Thioglobus',
        'CaAtelocyanobacterium' : 'Candidatus Atelocyanobacterium',
        '0319-6G20': 'Bdellovibrionales',
        'Euryarcheota' : 'Euryarchaeota' ,
        'SBR1093' : 'Bacteria',
        'Euryarcheoata' : 'Euryarchaeota'
        }

        regions = { 'NP' : 'North_Pacific',
        'NAT' : 'North_Atlantic',
        'MED' : 'Mediterranean',
        'ARS' : 'Arabian_Sea',
        'RS'  : 'Red_Sea',
        'IN'  : 'Indian_Ocean',
        'EAC' : 'East_Africa_Coastal',
        'SAT' : 'South_Atlantic',
        'CPC' : 'Chile_Peru_Coastal',
        'SP'  : 'South_Pacific'
        }

        wb2 = load_workbook("metadata/Table4_Phylogeny.xlsx")
        taxos = { l[0] : [v for v in l[:-1] if v != 'null' and not v[0:4] == "nove" ][-1] for l in wb2.get_sheet_by_name('Hug set').values}
        taxos = {k : corrected[v] if corrected.has_key(v) else v for k, v in taxos.items()}

        tax_2_id = self.taxDb.get_name_translator(taxos.values())
        tax_ids = {g : tax_2_id.get(taxos[g])[0]  for g in t_metadata['Genome ID'] if taxos.has_key(g) }
        t_metadata['species_taxid'] = [ tax_ids[g] if tax_ids.has_key(g) else 131567 for g in t_metadata['Genome ID']]
        t_metadata.index = Index(t_metadata['Genome ID'])
        t_metadata['region'] = [regions[g.split("_")[1].split("-")[0]] for g in t_metadata['Genome ID']]
        self.metadata = t_metadata.transpose().to_dict()

        print "Loading genomes"
        if os.path.exists(pjoin(self.data_path , 'TOBGGENOMES.tar.gz')):
            os.system("tar xzvf " + pjoin(self.data_path , 'TOBGGENOMES.tar.gz'))
            os.remove(pjoin(self.data_path , 'TOBGGENOMES.tar.gz'))

        for k,v in tqdm(self.metadata.items()):
            genome_path = pjoin(self.data_path, v['region'], k)
            genome_file = pjoin(genome_path, k + ".fna")
            if not os.path.exists(genome_file):
                os.makedirs(pjoin(genome_path, 'original_files'))
                shutil.move(self.data_path + k + ".fna", pjoin(genome_path, 'original_files'))
            self.genomes += [Genome(k, genome_path, ref=pjoin(genome_path, 'original_files', k + ".fna"), manual_metadata = v, taxDb = self.taxDb, workbench = self.workbench)]

    def process(self, num_cores = 10):
        Database.process(self)
